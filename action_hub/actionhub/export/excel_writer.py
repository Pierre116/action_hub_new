from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from actionhub.actions.queries import list_actions


def build_actions_workbook(filters: dict) -> BytesIO:
    result = list_actions({**filters, "page": 1, "per_page": 5000})
    rows = result["items"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Actions"

    headers = [
        "Ref / 编号",
        "Title / 标题",
        "Topic / 主题",
        "Secondary Topic / 第二主题",
        "Team / 团队",
        "Priority / 优先级",
        "Status / 状态",
        "Deadline / 截止日期",
        "Actual Date / 完成日期",
        "Created At / 创建时间",
    ]
    sheet.append(headers)

    for idx, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)

    overdue_fill = PatternFill(fill_type="solid", fgColor="FDECEA")
    for row_idx, item in enumerate(rows, start=2):
        sheet.append(
            [
                item.get("act_ref"),
                item.get("act_title"),
                item.get("topic_name"),
                item.get("secondary_topic_name"),
                item.get("team_name"),
                item.get("act_priority"),
                item.get("act_status"),
                item.get("act_deadline"),
                item.get("act_actual_date"),
                item.get("act_created_at"),
            ]
        )

        deadline = item.get("act_deadline")
        status = item.get("act_status")
        if deadline and status not in {"Done", "Cancelled"}:
            try:
                if datetime.fromisoformat(str(deadline)).date() < datetime.now(timezone.utc).date():
                    for col in range(1, len(headers) + 1):
                        sheet.cell(row=row_idx, column=col).fill = overdue_fill
            except ValueError:
                pass

    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
        sheet.column_dimensions[col].width = 22

    sheet.auto_filter.ref = f"A1:J{max(2, len(rows) + 1)}"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream

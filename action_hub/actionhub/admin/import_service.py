from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from actionhub.actions.service import get_initial_assignment_status
from actionhub.middleware.db import get_db
from actionhub.utils.ref_generator import generate_action_ref


IMPORT_CACHE: dict[str, dict] = {}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _find_header_row(sheet) -> tuple[int, list[str]]:
    for row_index in range(1, min(10, sheet.max_row) + 1):
        values = [_normalize_header(cell.value) for cell in sheet[row_index]]
        non_empty = [v for v in values if v]
        if len(non_empty) >= 2:
            return row_index, values
    return 1, [_normalize_header(cell.value) for cell in sheet[1]]


def _index_of(headers: list[str], aliases: list[str]) -> int | None:
    alias_set = {_normalize_header(a) for a in aliases}
    for idx, header in enumerate(headers):
        if header in alias_set:
            return idx
    for idx, header in enumerate(headers):
        if any(alias in header for alias in alias_set if alias):
            return idx
    return None


def _detect_format(sheet_name: str, headers: list[str]) -> str:
    head = " ".join(headers)
    if "action id" in head or "reference" in head or "act_ref" in head:
        return "v4"
    if "owner" in head and "priority" in head and ("department" in head or "team" in head):
        return "v3"
    if "responsible" in head or "due date" in head:
        return "v2"
    if "action" in head or "title" in head or "事项" in head:
        return "v1"
    lowered_name = sheet_name.lower()
    if "v4" in lowered_name:
        return "v4"
    if "v3" in lowered_name:
        return "v3"
    if "v2" in lowered_name:
        return "v2"
    return "v1"


def _resolve_user_id(raw_owner: str | None, owner_map: dict[str, int]) -> int | None:
    if not raw_owner:
        return None
    key = str(raw_owner).strip()
    if not key:
        return None
    if key in owner_map:
        return int(owner_map[key])

    db = get_db()
    row = db.execute(
        """
        SELECT usr_id
        FROM t_user
        WHERE usr_display_name = ? OR COALESCE(usr_display_name_cn, '') = ? OR usr_employee_id = ?
        LIMIT 1
        """,
        (key, key, key),
    ).fetchone()
    return int(row["usr_id"]) if row else None


def _resolve_team_id(raw_team: str | None, team_map: dict[str, int]) -> int | None:
    if not raw_team:
        return None
    key = str(raw_team).strip()
    if not key:
        return None
    if key in team_map:
        return int(team_map[key])

    db = get_db()
    row = db.execute(
        """
        SELECT tea_id
        FROM t_team
        WHERE tea_code = ? OR tea_name_en = ? OR COALESCE(tea_name_cn, '') = ?
        LIMIT 1
        """,
        (key, key, key),
    ).fetchone()
    return int(row["tea_id"]) if row else None


def _normalize_priority(raw_priority: object) -> str:
    text = str(raw_priority or "").strip().lower()
    mapping = {
        "critical": "Critical",
        "high": "High",
        "medium": "Medium",
        "low": "Low",
        "紧急": "Critical",
        "高": "High",
        "中": "Medium",
        "低": "Low",
    }
    return mapping.get(text, "Medium")


def _normalize_status(raw_status: object) -> str:
    text = str(raw_status or "").strip().lower()
    mapping = {
        "open": "Open",
        "in progress": "In Progress",
        "on hold": "On Hold",
        "done": "Done",
        "cancelled": "Cancelled",
        "未开始": "Open",
        "进行中": "In Progress",
        "暂停": "On Hold",
        "完成": "Done",
        "取消": "Cancelled",
    }
    return mapping.get(text, "Open")


def _normalize_deadline(raw_deadline: object) -> str | None:
    if raw_deadline is None or str(raw_deadline).strip() == "":
        return None
    if isinstance(raw_deadline, datetime):
        return raw_deadline.date().isoformat()
    text = str(raw_deadline).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return None


def preview_import(file_bytes: bytes, filename: str) -> dict:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    header_row_idx, headers = _find_header_row(sheet)
    detected_format = _detect_format(sheet.title, headers)

    title_idx = _index_of(headers, ["title", "action", "action item", "事项", "问题", "problem", "subject"])
    desc_idx = _index_of(headers, ["description", "desc", "details", "comment", "remarks", "说明"])
    owner_idx = _index_of(headers, ["owner", "lead", "responsible", "assignee", "负责人"])
    dept_idx = _index_of(headers, ["department", "dept", "team", "部门", "团队"])
    priority_idx = _index_of(headers, ["priority", "severity", "优先级"])
    status_idx = _index_of(headers, ["status", "state", "状态"])
    deadline_idx = _index_of(headers, ["deadline", "due", "due date", "target date", "截止", "完成日期"])

    if title_idx is None:
        raise ValueError("Unable to detect title column in import file")

    rows: list[dict] = []
    unresolved_owners: set[str] = set()
    unresolved_teams: set[str] = set()

    for excel_row in range(header_row_idx + 1, sheet.max_row + 1):
        values = [sheet.cell(row=excel_row, column=col).value for col in range(1, len(headers) + 1)]
        title = str(values[title_idx] or "").strip()
        if not title:
            continue

        item = {
            "row": excel_row,
            "title": title,
            "description": str(values[desc_idx] or "").strip() if desc_idx is not None else "",
            "owner": str(values[owner_idx] or "").strip() if owner_idx is not None else "",
            "team": str(values[dept_idx] or "").strip() if dept_idx is not None else "",
            "priority": _normalize_priority(values[priority_idx] if priority_idx is not None else None),
            "status": _normalize_status(values[status_idx] if status_idx is not None else None),
            "deadline": _normalize_deadline(values[deadline_idx] if deadline_idx is not None else None),
        }

        if item["owner"] and _resolve_user_id(item["owner"], {}) is None:
            unresolved_owners.add(item["owner"])
        if item["team"] and _resolve_team_id(item["team"], {}) is None:
            unresolved_teams.add(item["team"])

        rows.append(item)

    token = uuid4().hex
    IMPORT_CACHE[token] = {
        "filename": filename,
        "format": detected_format,
        "rows": rows,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    return {
        "token": token,
        "filename": filename,
        "format": detected_format,
        "total_rows": len(rows),
        "preview_rows": rows[:20],
        "unresolved_owners": sorted(unresolved_owners),
        "unresolved_teams": sorted(unresolved_teams),
    }


def execute_import(token: str, owner_map: dict[str, int], team_map: dict[str, int], actor_user_id: int, skip_duplicates: bool = True) -> dict:
    cached = IMPORT_CACHE.get(token)
    if not cached:
        raise ValueError("invalid or expired import token")

    rows = cached["rows"]
    db = get_db()

    cursor = db.execute(
        """
        INSERT INTO t_import_log (iml_filename, iml_total_rows, iml_imported_by, iml_status)
        VALUES (?, ?, ?, 'Completed')
        """,
        (cached["filename"], len(rows), actor_user_id),
    )
    import_log_id = int(cursor.lastrowid)

    imported = 0
    skipped = 0
    duplicates = 0
    warnings = 0
    warning_messages: list[str] = []

    for item in rows:
        title = str(item.get("title") or "").strip()
        if len(title) < 5:
            skipped += 1
            warnings += 1
            warning_messages.append(f"Row {item['row']}: title too short")
            continue

        team_id = _resolve_team_id(item.get("team"), team_map)

        owner_id = _resolve_user_id(item.get("owner"), owner_map)
        if owner_id is None:
            owner_id = actor_user_id
            warnings += 1
            warning_messages.append(f"Row {item['row']}: unresolved owner '{item.get('owner')}', defaulted to importer")

        duplicate = db.execute(
            "SELECT act_id FROM t_action WHERE act_title = ? AND act_team_id = ? LIMIT 1",
            (title, team_id),
        ).fetchone()
        if duplicate and skip_duplicates:
            duplicates += 1
            skipped += 1
            continue

        action_cursor = db.execute(
            """
            INSERT INTO t_action (
                act_ref, act_title, act_desc, act_team_id, act_priority, act_status,
                act_deadline, act_source, act_source_ref, act_created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'Import', ?, ?)
            """,
            (
                generate_action_ref(),
                title,
                item.get("description") or None,
                team_id,
                _normalize_priority(item.get("priority")),
                _normalize_status(item.get("status")),
                _normalize_deadline(item.get("deadline")),
                str(import_log_id),
                actor_user_id,
            ),
        )
        action_id = int(action_cursor.lastrowid)

        db.execute(
            """
            INSERT INTO t_assignment (asg_action_id, asg_user_id, asg_role, asg_status, asg_assigned_by)
            VALUES (?, ?, 'Lead', ?, ?)
            """,
            (action_id, owner_id, get_initial_assignment_status(), actor_user_id),
        )
        db.execute(
            """
            INSERT INTO t_action_history (ahi_action_id, ahi_change_type, ahi_field, ahi_new_value, ahi_changed_by)
            VALUES (?, 'Created', 'import', ?, ?)
            """,
            (action_id, f"import_log:{import_log_id}", actor_user_id),
        )
        imported += 1

    db.execute(
        """
        UPDATE t_import_log
        SET iml_imported = ?,
            iml_skipped = ?,
            iml_duplicates = ?,
            iml_warnings = ?,
            iml_warn_details = ?,
            iml_status = 'Completed'
        WHERE iml_id = ?
        """,
        (imported, skipped, duplicates, warnings, json.dumps(warning_messages, ensure_ascii=False), import_log_id),
    )
    db.commit()
    IMPORT_CACHE.pop(token, None)

    return {
        "import_log_id": import_log_id,
        "filename": cached["filename"],
        "total_rows": len(rows),
        "imported": imported,
        "skipped": skipped,
        "duplicates": duplicates,
        "warnings": warnings,
    }


def list_import_history() -> list[dict]:
    db = get_db()
    rows = db.execute(
        """
        SELECT iml_id, iml_filename, iml_total_rows, iml_imported, iml_skipped, iml_duplicates,
               iml_warnings, iml_status, iml_imported_at
        FROM t_import_log
        ORDER BY iml_imported_at DESC, iml_id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def rollback_import(import_log_id: int) -> dict:
    db = get_db()
    log = db.execute("SELECT iml_id, iml_status FROM t_import_log WHERE iml_id = ?", (import_log_id,)).fetchone()
    if not log:
        raise ValueError("import log not found")
    if log["iml_status"] == "Rolled Back":
        return {"import_log_id": import_log_id, "deleted_actions": 0, "status": "Rolled Back"}

    count_row = db.execute(
        "SELECT COUNT(*) AS total FROM t_action WHERE act_source = 'Import' AND act_source_ref = ?",
        (str(import_log_id),),
    ).fetchone()
    deleted_count = int(count_row["total"] or 0)

    db.execute(
        "DELETE FROM t_action WHERE act_source = 'Import' AND act_source_ref = ?",
        (str(import_log_id),),
    )
    db.execute(
        "UPDATE t_import_log SET iml_status = 'Rolled Back' WHERE iml_id = ?",
        (import_log_id,),
    )
    db.commit()

    return {"import_log_id": import_log_id, "deleted_actions": deleted_count, "status": "Rolled Back"}
